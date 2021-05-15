import openpyxl


# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook('transactions.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet1']


# wb.create_sheet()
# wb.remove_sheet()

cell = sheet['a1']
print(cell.value)
print(cell.row)
print(cell.column)
print(cell.coordinate)


sheet.cell['a1']
